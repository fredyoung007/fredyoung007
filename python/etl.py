#!/usr/bin/env python3
"""etl.py
Validate reference numbers with structure: da-dc-lc-gc
Seperator: -

valid: no fill
check_valid: light green, vcomment = "Please verify the Tracking Code"
partial_valide: yellow, pvcomment = "Invalid Tracking Code"
invalid: red
daInvalid: red, daicomment = "{da} is invalid abbreviation"
dcInvalid: red, dcicomment = "{dc} is invalid department code"
lcInvalid: red, gcicomment = "{lc} is invalid geo code"
"""
import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill

xfile = openpyxl.load_workbook("Valid financial codes list.xlsx")
sheetToVerify = xfile["To Verify"]
sheetValidCodes = xfile["Valid Codes"]
sheetVerified = xfile.create_sheet("Verified")
xfile.active = 2
da_range = sheetValidCodes["A:H"]
gc_range = sheetValidCodes["J:X"]

dals = ["CR", "EM", "FD", "HW", "IO", "PS", "RR", "SU"]
lcls = ["AB", "BC", "MB", "NB","NL", "NS", "NU", "NT", "ON", "PE", "QC", "SK", "Yk","NO","IN"]
costCode = "EGBCGRT"

# Validation status code
VALID_STATUS = [
    ("DA_INVALID", -1),
    ("DC_INVALID", -2),
    ("LC_INVALID", -3),
    ("CHECK_VALID", 1),
    ("PARTIAL_VALID", 2),
    ("VALID", 3)
]
DA_INVALID = -1
DC_INVALID = -2
LC_INVALID = -3
CHECK_VALID = 1
PARTIAL_VALID = 2
VALID = 3

# Status message
daiMsg = " is invalid abbreviation"
dciMsg = " is invalid department code"
lciMsg = " is invalid geo location"
gciMsg = " is invalid geo code"
cvMsg = "Please verify the Tracking Code"
pvMsg = "Invalid Tracking Code"

# Color
LIGHT_GREEN = "90ee90"
YELLOW = "ffff00"
RED = "ff0000"
WHITE = "ffffff"

# Test data
rn1 = "SU-851-NO-GEOCODE" # valid
rn2 = "EM-355-ON-ONNO" # valid
rn3 = "570-NO-CRPPLAN" # invalid
rn4 = "CR-570-NO-CRLEGRI" # check_valid
rn5 = "EM-340-BC-EGBCGRT" # partial_valid
rn6 = "ES-365-ON-OCCD" # da invalid
rn7 = "EM-900-ON-OCCD" # dc invalid
rn8 = "FD-201-AB-OCCD" # partial valid
rn9 = "CR-599-IN-ANCC" # partial valid
rn10 = "CR-575-IN" # partial valid

def loadDC(da_range):
    dicDC = {}
    for row in da_range:
        dcls = []
        for cell in row:
            if cell.value is not None:
                if cell.row == 1:
                    key = cell.value
                else:
                    dcls.append(str(cell.value))

        dicDC.update({key:dcls})
    return dicDC

def loadGC(gc_range):
    dicGC = {}
    for row in gc_range:
        gcls = []
        for cell in row:
            if cell.value is not None:
                if cell.row == 1:
                    key = cell.value
                else:
                    gcls.append(str(cell.value))

        dicGC.update({key:gcls})
    return dicGC

def loadToVerify(sheetToVerify):
    toVerify = []
    colA = sheetToVerify['A']
    for row in colA:
        if (row.value != None) and (row.row != 1):
            toVerify.append(row.value)
    return toVerify

def validateRefNumber(rn, dicDC, dicGC):
    errorCode = 0
    rnlst = rn.split("-")
    rnlength = len(rnlst)
    
    da = dc = lc = gc = ''

    if rnlength > 0:
        da = rnlst[0]
    if rnlength > 1:
        dc = rnlst[1]
    if rnlength > 2:
        lc = rnlst[2]
    if rnlength > 3:
        gc = rnlst[3]

    if da in dals:
        if dc in dicDC[da]:
            if lc == "NO":
                return [CHECK_VALID, cvMsg]
            elif lc in lcls:
                if lc is None:
                    return [VALID]
                if gc == costCode:
                    return [PARTIAL_VALID, pvMsg]
                if gc in dicGC[lc]:
                    return [VALID]
                else:
                    return [PARTIAL_VALID, pvMsg]
            else:
                return [LC_INVALID, lc + lciMsg]
        else:
            return [DC_INVALID, dc + dciMsg]
    else:
        return [DA_INVALID, da + daiMsg]

dicDC = loadDC(da_range)
dicGC = loadGC(gc_range)
sheetVerified["A1"].value = "Reference Number"

i = 2
fillColor = WHITE
for rn in loadToVerify(sheetToVerify):
    vc = validateRefNumber(rn, dicDC, dicGC)
    sheetVerified['A{}'.format(i)].value = rn
    if len(vc) == 2:
        sheetVerified['A{}'.format(i)].comment = Comment(vc[1], "FY")
        if vc[0] < 0:
            fillColor = RED 
        elif vc[0] == CHECK_VALID:
            fillColor = LIGHT_GREEN
        elif vc[0] == PARTIAL_VALID:
            fillColor = YELLOW 
        sheetVerified['A{}'.format(i)].fill = PatternFill("solid", fgColor=fillColor)
    i +=1

xfile.save("Valid financial codes list Verified.xlsx")


